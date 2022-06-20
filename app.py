import openpyxl

# wb = openpyxl.Workbook() #create empty workbook
wb = openpyxl.load_workbook("transactions.xlsx")

# print(wb.sheetnames) #['Sheet1']

sheet = wb["Sheet1"]  # create sheet object of Sheet1

# wb.create_sheet("Sheet0", 0)
# print(wb.sheetnames)  # ['Sheet0', 'Sheet1']

# wb.remove_sheet(wb["Sheet0"])
# print(wb.sheetnames)  # ['Sheet1']

cell = sheet["a1"]
cell = sheet.cell(row=1, column=1)  # they're the same

#cell = "helloworld"

print(cell.value, cell.column, cell.row, cell.coordinate)
# transaction_id 1 1 A1

print(sheet.max_row, sheet.max_column)  # 4 3

# iterate over whole sheet and print values in a nice way
# for row in range(sheet.max_row):
#     for column in range(sheet.max_column):
#         cell = sheet.cell(row=row+1, column=column+1)
#         print("{:^15s}".format(str(cell.value)), end="")
#     print("")

# columna_a = sheet['a'] #this will make a tuple of cell object of all cells in column a
# for cell in columna_a:
#     print(cell.value)

# # returns tuple of columns from a to c each in separte tuple
# columns = sheet["a:c"]
# for column in columns:
#     for cell in column:
#         print(cell.value)

# cells = sheet["a1:c3"]#returns tuple of rows(from 1 to 3) each row has a,b,c
# for row in cells:
#     for cell in row:
#         print(cell.value)

# rows = sheet[1:3]#returns tuple of rows(from 1 to 3) each row has a,b,c
# for row in rows:
#     for cell in row:
#         print(cell.value)

sheet.append([1, 2, 3])  # add row with the list of columns

wb.save('test.xlsx')
